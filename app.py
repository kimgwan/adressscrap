from flask import Flask, render_template, request, send_file, redirect, url_for, session, send_from_directory
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import re
from collections import defaultdict
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
from datetime import datetime, timedelta
import uuid
import threading

app = Flask(__name__)
app.secret_key = "your_secret_key_here"  # 세션을 위한 비밀 키
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=1)  # 세션 유효 시간 설정

# 파일 저장 디렉토리
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(UPLOAD_FOLDER, f'scraper_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'))
    ]
)
logger = logging.getLogger(__name__)

# 작업 상태 저장
task_status = {}

def get_hospital_data(name, password, taskID, user_id):
    """
    웹사이트에서 병원 정보를 스크래핑하고 엑셀 파일로 저장합니다.
    
    Args:
        name (str): 로그인 아이디
        password (str): 로그인 비밀번호
        taskID (str): 작업 ID
        user_id (str): 사용자 고유 ID
        
    Returns:
        str: 저장된 파일 경로
    """
    try:
        task_status[user_id] = {"status": "processing", "message": "작업이 시작되었습니다."}
        
        url = "https://agent-front.green-ribbon.co.kr/"
        if not taskID:
            task_status[user_id] = {"status": "error", "message": "업무무ID가 작업을 진행할 수 없습니다."}
            return None

        url2 = f"https://agent-front.green-ribbon.co.kr/v2/check-list/{taskID}"
        
        # 브라우저 옵션 설정 (헤드리스 모드)
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        
        driver = webdriver.Chrome(options=options)
        
        task_status[user_id] = {"status": "processing", "message": "웹사이트에 접속 중..."}

        # 1. 로그인
        driver.get(url)
        logger.info('웹사이트 접속')

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "name")))
        driver.find_element(By.NAME, "name").send_keys(name)
        logger.info('아이디 입력')
        
        driver.find_element(By.XPATH, '//*[@type="password"]').send_keys(password)
        logger.info('패스워드 입력')
        
        driver.find_element(By.XPATH, '//*[@type="submit"]').click()
        logger.info('로그인 완료')
        
        task_status[user_id] = {"status": "processing", "message": "로그인 완료, 데이터 수집 중..."}

        # 2. 작업 페이지 접근
        driver.get(url2)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body > main")))
        logger.info('main 태그 로딩 완료')

        # 3. 병원 데이터 추출
        block_elements = driver.find_elements(By.CSS_SELECTOR, "body > main > div > div.pt-2 > div")
        name_pattern = re.compile(r'^[가-힣]{2,4}$')
        exclude_keywords = {"고객정보복사", "실손", "특급대행", "진행중", "자보", "골절", "배정완료"}

        hospital_data = defaultdict(lambda: {"addresses": set(), "names": set()})
        
        task_status[user_id] = {"status": "processing", "message": f"총 {len(block_elements)}개 병원 정보 수집 중..."}

        for idx, block in enumerate(block_elements, 1):
            try:
                # 병원 이름 추출
                hospital_name = block.find_element(By.CSS_SELECTOR, "div.flex.items-center.space-x-2 > p.font-semibold").text
                
                # 주소 추출
                address = block.find_element(By.CSS_SELECTOR, "p.text-sm.underline.cursor-pointer").text
                
                # 토글 버튼 클릭
                toggle_button = block.find_element(By.CSS_SELECTOR, "div.flex.items-center.justify-center.border.border-input.py-2.rounded-md.mt-6.cursor-pointer")
                driver.execute_script("arguments[0].scrollIntoView(true);", toggle_button)
                
                # 토글 버튼 클릭 대기 및 실행
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 
                    "div.flex.items-center.justify-center.border.border-input.py-2.rounded-md.mt-6.cursor-pointer")))
                ActionChains(driver).move_to_element(toggle_button).click().perform()
                
                # 클릭 후 요소 로딩 대기
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "body > main > div > div.pt-2 > div"))
                )
                
                # 이름 요소 추출
                p_elements_name = block.find_elements(By.CSS_SELECTOR, "div p")
                
                # 데이터 저장
                hospital_data[hospital_name]["addresses"].add(address)
                for p in p_elements_name:
                    text = p.text.strip()
                    if name_pattern.match(text) and text not in exclude_keywords:
                        hospital_data[hospital_name]["names"].add(text)
                        
                # 상태 업데이트
                if idx % 5 == 0:
                    task_status[user_id] = {"status": "processing", "message": f"{idx}/{len(block_elements)} 병원 정보 수집 중..."}
                    
            except Exception as e:
                logger.error(f"오류 발생 (블록 {idx} 처리 중): {e}")

        task_status[user_id] = {"status": "processing", "message": "데이터 수집 완료, 엑셀 파일 생성 중..."}

        # 4. 엑셀 데이터 준비
        excel_data = {"병원 이름": [], "주소": [], "추출된 이름": []}
        for hospital_name, data in hospital_data.items():
            excel_data["병원 이름"].append(hospital_name)
            excel_data["주소"].append(", ".join(data['addresses']))
            excel_data["추출된 이름"].append(", ".join(data['names']) if data['names'] else "없음")

        # 5. pandas DataFrame 생성
        df = pd.DataFrame(excel_data)

        # 6. 주소 처리 함수
        def process_address(address):
            # 원본 주소 저장
            original_address = address
            
            # 공백 제거
            address = address.replace(" ", "")
            
            # 쉼표 이후의 부분 제거 (첫 번째 쉼표만)
            comma_idx = address.find(",")
            if comma_idx != -1:
                paren_idx = address.find("(", comma_idx)
                if paren_idx != -1:
                    address = address[:comma_idx] + address[paren_idx:]
                else:
                    address = address[:comma_idx]
            
            import re
            
            # 특수 케이스 확인: "대원로8 2,3층" 같은 패턴 (숫자 다음에 층이 바로 오는지)
            special_pattern = re.search(r'([가-힣]+로|[가-힣]+길)([0-9-]+)([0-9]+층)', original_address)
            if special_pattern:
                # 스페이스나 쉼표로 구분된 경우 별도 처리
                if ' ' in original_address or ',' in original_address:
                    # 도로명과 번호까지만 추출
                    road_name = special_pattern.group(1)
                    road_num = special_pattern.group(2)
                    
                    # 괄호 부분 추출
                    paren_match = re.search(r'\([^)]+\)', original_address)
                    paren_part = paren_match.group(0) if paren_match else ""
                    
                    return road_name + road_num + paren_part
            
            # 괄호 부분 임시 저장
            paren_parts = []
            for match in re.finditer(r'\([^)]+\)', address):
                paren_parts.append(match.group())
            
            # 괄호 제외한 부분만 남기기
            clean_addr = re.sub(r'\([^)]+\)', '§', address)
            
            # 도로명+번호 패턴 찾기
            road_num_pattern = re.search(r'([가-힣]+로|[가-힣]+길)([0-9-]+)', clean_addr)
            if road_num_pattern:
                road_end_idx = road_num_pattern.end()
                # 도로명+번호 이후 부분
                after_road = clean_addr[road_end_idx:]
                
                # 층, 호수 등 불필요한 정보 제거
                clean_after = re.sub(r'[0-9]+층|[0-9]+호|[A-Za-z]동|[가-힣]동', '', after_road)
                
                # 괄호 직전까지의 텍스트 (건물명 등) 제거
                next_paren_idx = clean_after.find('§')
                if next_paren_idx > 0:  # 괄호 앞에 텍스트가 있으면
                    clean_after = clean_after[next_paren_idx:]
                
                clean_addr = clean_addr[:road_end_idx] + clean_after
            
            # 괄호 원복
            paren_idx = 0
            while '§' in clean_addr and paren_idx < len(paren_parts):
                clean_addr = clean_addr.replace('§', paren_parts[paren_idx], 1)
                paren_idx += 1
            
            return clean_addr

        # 7. 주소 열에 처리 적용
        df["주소"] = df["주소"].apply(process_address)

        # 8. 주소를 기준으로 오름차순 정렬
        df = df.sort_values(by="주소", ascending=True)

        # 9. 엑셀 파일 저장을 위한 사용자별 디렉토리 생성
        user_dir = os.path.join(UPLOAD_FOLDER, user_id)
        if not os.path.exists(user_dir):
            os.makedirs(user_dir)

        # 10. 엑셀 파일 저장
        current_datetime = time.strftime("%Y%m%d_%H%M%S")
        excel_filename = f"hospital_data_{current_datetime}.xlsx"
        excel_path = os.path.join(user_dir, excel_filename)
        df.to_excel(excel_path, index=False, engine='openpyxl')

        # 11. openpyxl로 파일 열고 중복된 주소에 색상 적용
        wb = load_workbook(excel_path)
        ws = wb.active

        # 노란색 배경 스타일 정의
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 주소 열(B열, 인덱스 2)에서 중복된 주소를 찾기
        address_col = 2
        address_count = defaultdict(int)
        for row in range(2, ws.max_row + 1):  # 헤더 제외, 2행부터
            address = ws.cell(row=row, column=address_col).value
            address_count[address] += 1

        # 중복된 주소에 노란색 적용
        for row in range(2, ws.max_row + 1):
            address = ws.cell(row=row, column=address_col).value
            if address_count[address] > 1:  # 중복된 경우
                ws.cell(row=row, column=address_col).fill = yellow_fill

        # 12. 수정된 엑셀 파일 저장
        wb.save(excel_path)
        logger.info(f"엑셀 파일 저장 완료 (중복 주소 색상 적용): {excel_path}")

        # 13. HTML 페이지 소스 저장
        html_path = os.path.join(user_dir, 'page_source.html')
        with open(html_path, 'w', encoding='utf-8') as file:
            file.write(driver.page_source)
        logger.info(f"HTML 저장 경로: {html_path}")
        
        task_status[user_id] = {
            "status": "completed", 
            "message": "작업이 완료되었습니다.", 
            "file": excel_filename,
            "user_id": user_id  # 사용자 ID도 저장
        }
        return excel_filename

    except Exception as e:
        logger.error(f"오류 발생: {e}")
        task_status[user_id] = {"status": "error", "message": f"오류 발생: {str(e)}"}
        return None

    finally:
        if 'driver' in locals():
            driver.quit()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    # 사용자 고유 ID 생성 (세션)
    if 'user_id' not in session:
        session['user_id'] = str(uuid.uuid4())
    
    return render_template('dashboard.html')

@app.route('/scrape', methods=['POST'])
def scrape():
    name = request.form.get('name')
    password = request.form.get('password')
    taskID = request.form.get('taskID')
    
    if not all([name, password, taskID]):
        return render_template('dashboard.html', error="모든 필드를 입력해주세요.")
    
    user_id = session.get('user_id', str(uuid.uuid4()))
    session['user_id'] = user_id  # 세션에 저장 확실히
    
    # 비동기로 작업 실행
    thread = threading.Thread(target=get_hospital_data, args=(name, password, taskID, user_id))
    thread.daemon = True
    thread.start()
    
    return redirect(url_for('status'))

@app.route('/status')
def status():
    user_id = session.get('user_id')
    if not user_id or user_id not in task_status:
        return render_template('status.html', status={"status": "not_started", "message": "작업이 시작되지 않았습니다."})
    
    return render_template('status.html', status=task_status[user_id])

@app.route('/downloads/<path:filename>')
def download_file(filename):
    user_id = session.get('user_id')
    if not user_id:
        return "사용자 정보를 찾을 수 없습니다. 다시 로그인해주세요.", 403
    
    user_dir = os.path.join(UPLOAD_FOLDER, user_id)
    if not os.path.exists(os.path.join(user_dir, filename)):
        return "파일을 찾을 수 없습니다.", 404
    
    return send_from_directory(user_dir, filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)